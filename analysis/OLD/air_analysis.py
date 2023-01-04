import os
import numpy as np
import scipy.ndimage as scind
from PIL import Image
import operator
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from scipy import stats
from sklearn import preprocessing

from utils_analysis_new import fit_circle, initialize_data


class air_analysis:
    def __init__(self, settings, metadata):

        # Step 1: initialize data and dictionary
        init = initialize_data(settings, metadata)
        init.run()

        self.meta_data = init.meta_data
        self.params = init.params
        self.label = init.label
        self.im = init.im

        # Set path for saving data
        self.path = init.analysis_root
        self.path_save_excel = os.path.join(self.path, self.params['path_save'])
        self.path_save_images = os.path.join(self.path, self.params['path_save'], self.label)
        os.makedirs(self.path_save_images, exist_ok = True)

        # Create dictionary for results report
        self.report = {}

        # Create variables for analysis
        self.mask_rev = None
        self.is_curved = False

    def remove_small_cluster(self):

        # Determine size of image
        h, w = np.shape(self.cca)

        # Determine size of minimum size of cluster using parameter clusters
        min_size = h * w * self.params['cluster_fminsize']

        # Calculate the clustersizes present in the image
        cluster_size = scind.sum(self.BW, self.cca, range(self.labels + 1))

        # Check if minimal cluster size is correct, if too large, divide by 10
        while sum(cluster_size > min_size) < 2 and min_size > 100:
            min_size = int(min_size / 10)

        # Remove small clusters
        mask_size = cluster_size < min_size  # Determine the small clusters
        self.cca[mask_size[self.cca]] = 0  # Set small cluster to 0
        labels_uni = np.unique(self.cca)
        labels = len(labels_uni)
        cca_new = np.searchsorted(labels_uni, self.cca)  # Create new components array without small clusters

        # Define new components and labels with the small components removed
        self.cca = cca_new
        self.labels = labels

    def select_vertical_middle(self):
        # Determine size of image
        h, w = np.shape(self.cca)

        # Search for clusters present in vertical middle area of image
        search_middle = self.cca[:, int(0.4 * w):int(0.6 * w)]

        labels_middle = []
        for ss in search_middle.ravel():
            if ss > 0 and not ss in labels_middle:  # Find label of components present in vertical middle
                labels_middle.append(ss)  # Append to label

        return labels_middle

    def isolate_reverberation(self):
        """
        1. Restrict to bbox if provided
        2. Find reverberations as largest connected component
        3. Return reverb mask
        4. Save image and report results
        """

        # Initialize report section for pattern
        report_section = 'pattern'
        self.report[report_section] = {}

        '''1. Create binary image with provided signal threshold'''
        signal_thresh = self.params['signal_thresh']
        self.BW = self.im > signal_thresh

        '''2. Find transducer image as largest connected component'''
        # Determine components and their labels
        self.cca, self.labels = scind.label(self.BW)

        # Remove small clusters
        self.remove_small_cluster()

        # Search for clusters present in the vertical middle
        labels_middle = self.select_vertical_middle()

        '''3. Return transducer image mask'''
        # Create reverberation mask containing only clusters larger than minimum size and present in vertical middle
        self.transducer_mask = np.reshape(np.in1d(self.cca, labels_middle), np.shape(self.cca))

        '''3a. Delete colorbars from sides if initial values provided'''
        if not self.params['init_x0x1'] is None:
            x0,x1 = self.params['init_x0x1']      # Provided initial box
            self.transducer_mask[:,:x0] = False           # Select box from image
            self.transducer_mask[:,x1+1:] = False

        '''3b. Save transducer image and report results'''
        # Create and save image of only largest cluster
        trans_image = self.im * self.transducer_mask
        trans_image = Image.fromarray(trans_image)
        trans_image.save(f"{self.path_save_images}/{self.label}image_transducer.png")

        # Save upper and lower edges of the reverberation pattern
        # Create list with indices of the cluster representing the reverberation pattern
        clus = np.where(self.transducer_mask)
        clus = [(x, y) for x, y in zip(clus[0], clus[1])]

        # Determine upper and lower edges of the cluster
        trans_minx = min(clus, key=operator.itemgetter(1))[1]
        trans_maxx = max(clus, key=operator.itemgetter(1))[1]
        trans_maxy = max(clus, key=operator.itemgetter(0))[0]
        trans_miny = min(clus, key=operator.itemgetter(0))[0]

        # Save upper and lower edges
        self.params['trans_x0y0x1y1'] = [trans_minx, trans_miny, trans_maxx, trans_maxy]

        self.report[report_section]['trans_box_xmin_px'] = ('int', trans_minx)
        self.report[report_section]['trans_box_ymin_px'] = ('int', trans_miny)
        self.report[report_section]['trans_box_xmax_px'] = ('int', trans_maxx)
        self.report[report_section]['trans_box_ymax_px'] = ('int', trans_maxy)

        '''4. Return reverberation image mask by selecting ROI'''
        self.reverb_mask = self.transducer_mask.copy()

        if not self.params['init_y0y1'] is None:
            y0,y1 = self.params['init_y0y1']      # Provided initial box
            self.reverb_mask[:y0,:] = False
            self.reverb_mask[y1+1:,:]  = False

        '''4a. Save image and report results'''
        # Create and save image of only largest cluster
        reverb_image = self.im * self.reverb_mask
        reverb_image = Image.fromarray(reverb_image)
        reverb_image.save(f"{self.path_save_images}/{self.label}image_reverberation.png")

        # Save upper and lower edges of the reverberation pattern
        # Create list with indices of the cluster representing the reverberation pattern
        clus = np.where(self.reverb_mask )
        clus = [(x, y) for x, y in zip(clus[0], clus[1])]

        # Determine upper and lower edges of the cluster
        rev_minx = min(clus, key=operator.itemgetter(1))[1]
        rev_maxx = max(clus, key=operator.itemgetter(1))[1]
        rev_maxy = max(clus, key=operator.itemgetter(0))[0]
        rev_miny = min(clus, key=operator.itemgetter(0))[0]

        # Save upper and lower edges
        self.params['reverb_x0y0x1y1'] = [rev_minx, rev_miny, rev_maxx, rev_maxy]

        self.report[report_section]['reverb_box_xmin_px'] = ('int', rev_minx)
        self.report[report_section]['reverb_box_ymin_px'] = ('int', rev_miny)
        self.report[report_section]['reverb_box_xmax_px'] = ('int', rev_maxx)
        self.report[report_section]['reverb_box_ymax_px'] = ('int', rev_maxy)


    def check_if_curved(self, reverberation_mask, type = 'reverberation'):
        """
        Function to determine if the image is acquired by a curved transducer

        1. Fit circle through top of reverberation image
        2. Determine if transducer is curved
        3. If the transducer is curved calculate limiting angles and radius of the fitted circle
        """
        report_section = 'curved'
        self.report[report_section] = {}

        ''' 1. Fit circle through top of reverberation image '''
        # Find outer edges of the image
        if type == 'reverberation':
            x0, y0, x1, y1 = self.params['reverb_x0y0x1y1']
        else:
            x0, y0, x1, y1 = self.params['trans_x0y0x1y1']

        # Find center x value
        center_x = int(0.5 * (x0 + x1) + .5)

        # From top of reverberation image down, look for pixels != 0 from mid to left and from mid to right
        # If both found, add it to the list
        circL_xy = []
        circR_xy = []
        for y in range(y0, y1 + 1):
            for x in reversed(range(x0, center_x)):  # find left point
                xl = -1
                xr = -1
                if reverberation_mask[y, x]:
                    xl = x
                    if xl > -1:
                        circL_xy.append((xl, y))
                    break
            for x in range(center_x, x1):  # find right point
                if reverberation_mask[y, x]:
                    xr = x
                    if xr > -1:
                        circR_xy.append((xr, y))
                    break

            if xr - xl < 10:            # Stop when
                break

        # Create complete circle to fit image
        circ_xy = []
        circ_xy.extend(circL_xy)
        circ_xy.extend(circR_xy)
        circ_xy.sort(key=operator.itemgetter(1))
        circle_fitfrac = self.params['circle_fitfrac']

        ''' 2.  Determine if transducer is curved '''
        if len(circ_xy) < 11:
            # at least 10 point, else probably not a curved probe
            is_curved = False

        else:
            is_curved = True

            '''3. Calculate limiting angles and radii '''
            # As deformation towards edges occur, use only central part for fitting
            fraction = 1 - circle_fitfrac
            cf_fraction = circ_xy[int(fraction * len(circ_xy)):]        # Create circle of only central part

            # Fit the fractionated circle: calculate optimized center of circle and determine radius of the circle
            cf = fit_circle(cf_fraction)
            (x_center, y_center), R = cf.circ_fit()

            # Calculate the limiting curvature of the circle at left side and right side in rad
            curve_left = np.arctan2(circL_xy[0][0] - x_center, circL_xy[0][1] - y_center)
            curve_right = np.arctan2(circR_xy[0][0] - x_center, circR_xy[0][1] - y_center)
            angles_curve = [curve_left, curve_right]

            # Calculate the limiting radii of the circle
            max_R = min([
                (x0 - x_center) / np.sin(angles_curve[0]),
                (x1 - x_center) / np.sin(angles_curve[1]),
                (y1 - y_center)])


            # Save and report the values
            self.params['curve_radii_px'] = [R, max_R]
            self.params['curve_origin_px'] = [x_center, y_center, R]
            self.params['curve_angles_deg'] = [c / np.pi * 180. for c in angles_curve]          # Save curved angles in degrees

            self.report[report_section]['curve_residue'] = ('float', cf.residue)                              # residue of fitted curvature
            self.report[report_section]['curve_radmin_px'] = ('float', self.params['curve_radii_px'][0])        # minimum radius of circle
            self.report[report_section]['curve_radmax_px'] = ('float', self.params['curve_radii_px'][1])        # maximum radius of circle
            self.report[report_section]['curve_xc_px'] = ('float', x_center)                                    # x center of circle
            self.report[report_section]['curve_yc_px'] = ('float', y_center)                                    # y center of circle
            self.report[report_section]['curve_rc_px'] = ('float', R)                                           # Radius of circle
            self.report[report_section]['curve_angmin_deg'] = ('float', self.params['curve_angles_deg'][0])     # minimum angle of circle curvature
            self.report[report_section]['curve_angmax_deg'] = ('float', self.params['curve_angles_deg'][1])     # maximum angle of circle curvature

        return is_curved

    def curve_straightening(self, reverberation_mask, type = 'reverberation'):
        """
        transform reverberation pattern to rectangle: interpolate at coords
        """
        # Create segmented image by using the reverberation mask
        im_segment = self.im * reverberation_mask

        # Define edges of the reverberation pattern
        if type == 'reverberation':
            x0, y0, x1, y1 = self.params['reverb_x0y0x1y1']
        else:
            x0, y0, x1, y1 = self.params['trans_x0y0x1y1']

        # Define the limiting angles and radius of the curvature of the pattern
        curve_angles = [c / 180. * np.pi for c in self.params['curve_angles_deg']]  # transform back to rad
        curve_radii = self.params['curve_radii_px']                                 # [radius, max_radius]

        # Define center values and radius of curvature
        curve_xyr = self.params['curve_origin_px']                                  # [x_center, y_center, radius]

        # Define vector containing all angle values between left limiting angle value and right limiting angle value
        angles = np.linspace(curve_angles[0], curve_angles[1], x1 - x0)

        # Define vector containing all radii values between minimum radius and maximum radius value
        radii = np.linspace(curve_radii[0], curve_radii[1], int(0.5 + curve_radii[1] - curve_radii[0]))

        # Create Cartesian matrix from coordinate vectors (angles & radii)
        an, ra = np.meshgrid(angles, radii)

        # Define x coordinates and y coordinates at which input (im_segment) is evaluated.
        xi = curve_xyr[0] + ra * np.sin(an)                 # Transform x values using x center as offset and meshgrid
        yi = curve_xyr[1] + ra * np.cos(an)                 # Transform y values using y center as offset and meshgrid

        # Create array of coordinates to find for each point in the output, the corresponding coordinates in the input
        coordinates = np.array([yi, xi])

        # Map the input array to new coordinates by interpolation.
        rect_image = scind.map_coordinates(im_segment, coordinates)

        # Create and save image of rectangular image
        image_rect = Image.fromarray(rect_image)
        image_rect.save(f"{self.path_save_images}/{self.label}image_uncurved_cropped.png")

        return rect_image

    def crop_image(self, reverberation_mask, type = 'reverberation'):
        """
        Crop the data to a box containing the reverberation pattern/ transducer image
        """

        # Create segmented image by using the reverberation mask
        im_segment = self.im * reverberation_mask

        # Define edges of the pattern
        if type == 'reverberation':
            x0, y0, x1, y1 = self.params['reverb_x0y0x1y1']
        else:
            x0, y0, x1, y1 = self.params['trans_x0y0x1y1']

        # Define pixels to skip for left and right (hcor_px) and above and below (vcor_px)
        dx = self.params['hcor_px']
        dy = self.params['vcor_px']

        # Crop the image to the edges of the reverberation pattern
        rect_image = im_segment[y0 + dy:y1 - dy + 1, x0 + dx:x1 - dx + 1]

        # Create and save image of cropped rectangular image
        image_rect = Image.fromarray(rect_image)
        image_rect.save(f"{self.path_save_images}/{self.label}image_cropped.png")

        return rect_image

    def pixels2mm(self, pixels):
        """
        translate pixels into mm
        """

        # Define physical delta x and physical unit in x direction
        delta_x = self.meta_data['Phys_delta_X']
        unit = self.meta_data['Phys_units_X']

        # Translate pixels into mm, but make sure unit is in mm
        if unit == 3:
            length_mm = 10. * delta_x * len(pixels)
        else:
            raise ValueError("Unit has to be set to mm")

        return length_mm

    def count_elements_neighboring(self, indices):
        """
        count number of elements in list and number of neighboring elements
        """

        neighbors = 0
        if len(indices) > 0:
            indices = sorted(indices)
            for i in range(len(indices) - 1):
                if indices[i + 1] == indices[i] + 1:
                    neighbors += 1

        return len(indices), neighbors

    def plot_stacked_pattern_profile(self, profile, im, mean, weak, dead, pixels10, pixels30, profile_type = 'horizontal'):

        x_values = np.array(list(range(len(profile))))

        if profile_type == 'vertical':
            im = np.transpose(im, (1,0))

        # Set up figure
        fig = plt.figure()

        # Set up axis; ax0 for uniformity image and ax1 for the horizontal profile
        ax0 = plt.axes([0.10, 0.76, 0.85, 0.20])
        ax1 = plt.axes([0.10, 0.10, 0.85, 0.65])

        # Show uniformity image on top and fill whole image
        ax0.imshow(im, cmap='gray', aspect='auto')

        # Make the tick labels invisible
        plt.setp(ax0.get_xticklabels(), visible=False)
        plt.setp(ax0.get_yticklabels(), visible=False)

        # Show profile below
        ax1.plot(x_values, profile, label="profile")

        # Determine x_max & y_max
        ymax = np.max(profile) + 0.1
        xmax = np.max(x_values)

        # Set the x and y labels
        ax1.set_xlabel("position [px]")
        ax1.set_ylabel("signal [a.u.]")
        ax1.grid(which='major', axis='y')
        ax1.legend()

        if profile_type == 'horizontal':
            # Plot the weak, dead and mean limits and plot the buckets
            ax1.plot([x_values[0], x_values[-1]], [mean, mean], linestyle=':', linewidth=2, color='green', label="average")
            ax1.plot([x_values[0], x_values[-1]], [weak, weak], linestyle=':', linewidth=2, color='orange', label="weak")
            ax1.plot([x_values[0], x_values[-1]], [dead, dead], linestyle=':', linewidth=2, color='red', label="dead")

            # Plot the buckets
            ax1.axvspan(x_values[0], x_values[pixels10], facecolor='black', alpha=0.2)      # Outer left 10 percent
            ax1.axvspan(x_values[-1], x_values[-pixels10], facecolor='black', alpha=0.2)    # Outer right 10 percent
            ax1.axvspan(x_values[pixels10], x_values[pixels30], facecolor='black', alpha=0.1) # Outer left 10-30 percent
            ax1.axvspan(x_values[-pixels30], x_values[-pixels10], facecolor='black', alpha=0.1) #Outer right 10-30 percent
            ax1.set_ylim(bottom=0, top=ymax)

            # Offset for xlimit to see the first and last lines
            xoffset = 5
            ax1.set_xlim(left=-xoffset, right=xmax + xoffset)
            ax0.set_xlim(left=-xoffset, right=xmax + xoffset)

            self.fig_uni = fig

            # Save image
            fig.savefig(f"{self.path_save_images}/{self.label}uniformity.png", bbox_inches='tight')

        else:
            # Set limits for x and y axis
            ax0.set_xlim(left=0, right=xmax)
            ax1.set_ylim(bottom=0, top=ymax)
            ax1.set_xlim(left=0, right=xmax)

            self.fig_sens = fig

            # Save image
            fig.savefig(f"{self.path_save_images}/{self.label}sensitivity.png", bbox_inches='tight')


    def calc_uniformity_parameters(self, u, buckets, report_section):
        """
        u_cov : scalar, horizontal profile covariance
        u_skew : scalar, horizontal profile skewness
        u_low : list, horizontal profile segment 10%, 20%, 40%, 20%, 10%
            MSE minimum for each segment"""

        # Evaluate the uniformity parameters from the horizontal profile (u)
        x_u = np.linspace(0, 100, len(u))

        # Calculate the coefficient of Variation over uniformity data = u_cov
        u_cov = 100 * (np.std(u) / np.mean(u))

        # Calculate the coefficient of skewness of the uniformity data = u_skew
        u_skew = stats.skew(u)

        # Analyse lowest value per bucket
        for lab, idx_bucket in buckets:
            v = u[idx_bucket]
            low_val = 100 * np.min((v - np.median(u)) / np.median(u))  # Calculate lowest value of segment

            # Report the calculated values
            self.report[report_section]['u_low_{}'.format(lab)] = ('int', low_val)

        #Report calculated values
        self.report[report_section]['u_cov'] = ('float', u_cov)
        self.report[report_section]['u_skew'] = ('float', u_skew)

    def normalize(self,data):
        return (data - data.min())/ (data.max() - data.min())

    def uniformity(self, im):
        """
        1. Average in vertical direction.
        2. Define threshold for weak and dead elements, and calculate line profiles for weak, dead and mean
        3. Count the number of pixel with response below the weak threshold, and the dead threshold.
            a) Separate analysis for all, the outer 10%, outer 10%-30%, and 30%-70%
            b) Also report the relative (to the overall mean) mean, min, max of all region
        4. Create image with reverberation pattern and plotted profile
        5. Extract uniformity parameters: u_cov, u_skew & u_low
        """

        report_section = "uniformity"
        self.report[report_section] = {}

        ''' 1. Average in vertical direction to create horizontal profile.'''
        horizontal_profile = np.average(im, axis=0)
        horizontal_profile = self.normalize(horizontal_profile)

        #  Report the width of the profile in pixels or degrees
        if self.is_curved == False:
            self.report[report_section]['unif_box_width_px'] = ('int', len(horizontal_profile))
        else:
            ang0, ang1 = self.params['curve_angles_deg']
            self.report[report_section]['unif_curve_width_deg'] = ('float', ang1 - ang0)

        # Report the width of the profile in mm
        length_mm = self.pixels2mm(horizontal_profile)
        self.report[report_section]['unif_width_mm'] = ('float', length_mm)

        '''2. Define threshold for weak and dead elements, and calculate line profiles for weak, dead and mean'''
        # Define threshold for weak and dead element values
        weak = self.params['f_weak']
        dead = self.params['f_dead']

        # Create line profiles for mean, weak and dead
        mean = np.average(horizontal_profile)
        weak = weak * mean
        dead = dead * mean

        '''3. Count the number of pixel with response below the weak threshold, and the dead threshold.'''
        # Find indices of pixels below the weak and dead thresholds
        weak_idx = sorted(np.where(horizontal_profile < weak)[0])
        dead_idx = sorted(np.where(horizontal_profile < dead)[0])

        # Count number of elements in list and number of neighboring elements
        elements = len(horizontal_profile)
        indices = list(range(elements))

        '''3a) Separate analysis for all, the outer 10%, outer 10%-30%, and 30%-70%'''
        pixels10 = max(0, min(int(elements * .1 + .5), elements - 1))
        pixels30 = max(0, min(int(elements * .3 + .5), elements - 1))

        # Create buckets to be analysed:
        buckets = [
            ('all', indices),                                                           # All indices
            ('outer10', indices[:pixels10] + indices[-pixels10:]),                      # Outer 10 precent of indices, left & right
            ('outer10_30', indices[pixels10:pixels30] + indices[-pixels30:-pixels10]),  # Outer 30 precent of indices, left & right
            ('inner10_90', indices[pixels10:-pixels10]),                                # Inner 10-90 precent of indices
            ('inner30_70', indices[pixels30:-pixels30]),                                # Inner 30-70 precent of indices
        ]

        # Analyse number of weak/dead elements and number of neighbours per bucket
        for lab, idx_bucket in buckets:
            # Calculate number of weak/dead elements in the bucket and number of neighboring elements
            weak_num, weak_neighbors = self.count_elements_neighboring([i for i in weak_idx if i in idx_bucket])
            dead_num, dead_neighbors = self.count_elements_neighboring([i for i in dead_idx if i in idx_bucket])

            # Report the calculated values
            self.report[report_section]['weak_{}'.format(lab)] = ('int', weak_num)
            self.report[report_section]['weakneighbors_{}'.format(lab)] = ('int', weak_neighbors)
            self.report[report_section]['dead_{}'.format(lab)] = ('int', dead_num)
            self.report[report_section]['deadneighbors_{}'.format(lab)] = ('int', dead_neighbors)

            '''b) Also report the relative (to the overall mean) mean, min, max of all region'''
            bucket_profile = horizontal_profile[idx_bucket]

            if not lab == 'all':
                self.report[report_section]['relmean_{}'.format(lab)] = ('float', np.average(bucket_profile) / mean)
                self.report[report_section]['relmin_{}'.format(lab)] = ('float', np.min(bucket_profile) / mean)
                self.report[report_section]['relmax_{}'.format(lab)] = ('float', np.max(bucket_profile) / mean)

        # Report other results
        self.report[report_section]['mean'] = ('float', mean)
        self.report[report_section]['relmin'] = ('float', np.min(horizontal_profile) / mean)
        self.report[report_section]['relmax'] = ('float', np.max(horizontal_profile) / mean)
        self.report[report_section]['f_weak'] = ('float', self.params['f_weak'])
        self.report[report_section]['f_dead'] = ('float', self.params['f_dead'])

        ''' 4. Create image with reverberation pattern and plotted profile '''
        self.plot_stacked_pattern_profile(horizontal_profile, im, mean, weak, dead, pixels10, pixels30, profile_type='horizontal')

        '''5. Extract uniformity parameters: u_cov, u_skew & u_low'''


    def crop_sens_image(self, im):
        """
        Take only middle 33% for sensitivity analysis
        """
        h, w = im.shape

        pixels_middle = max(0, min(int(w * .3 + .5), w - 1))
        im_sens = im[:, pixels_middle:-pixels_middle]

        sens_im = Image.fromarray(im_sens)
        sens_im.save(f"{self.path_save_images}/{self.label}image_cropped_sensitivity.png")

        return im_sens

    def sensitivity(self, im):
        """
        1. Average in horizontal direction to create vertical profile
        2. Report the maximum depth of the reverberation pattern
        3. Create image with reverberation pattern and plotted vertical profile
        """

        report_section = "sensitivity"
        self.report[report_section] = {}

        '''1. Average in horizontal direction to create vertical profile'''
        vertical_profile = np.average(im, axis=1)
        vertical_profile = self.normalize(vertical_profile)
        y_values = np.array(list(range(len(vertical_profile))))

        '''Report the maximum depth of the reverberation pattern'''
        # # Determine s_depth as the intersection of the noise and the vertical intensity profile
        # background = self.im[np.round(im.shape[0] / 2).astype(int):, :]
        # background_value = np.median(background.ravel())
        # s_depth = np.argmin(np.abs(vertical_profile - background_value))

        ''' 3. Create image with reverberation pattern and plotted vertical profile'''
        self.plot_stacked_pattern_profile(vertical_profile, im, None, None, None, None, None, profile_type='vertical')

    def export_results(self):

        if os.path.exists(f"{self.path_save_excel}/results.xlsx"):
            wb = load_workbook(f"{self.path_save_excel}/results.xlsx")
            if self.label in wb.sheetnames:
                wb.remove(wb[self.label])
                sheet = wb.create_sheet(title=self.label)
            else:
                sheet = wb.create_sheet(title=self.label)
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.title = self.label

        sheet["A1"] = "Section"
        sheet["B1"] = "Variable"
        sheet["C1"] = "Value"

        start_row = 2
        for section in self.report.keys():
            sheet[f"A{start_row}"] = section

            for row2, (key, vals) in enumerate(self.report[section].items()):
                row2 += start_row
                sheet[f"B{row2}"] = key
                sheet[f"C{row2}"] = vals[1]

            start_row += len(self.report[section])

        wb.save(f"{self.path_save_excel}/results.xlsx")

    def plot_final_fig(self):

        # Set up figure
        fig_uni = self.fig_uni
        fig_sens = self.fig_sens

        # # Set up axis; ax0 for uniformity image and ax1 for the horizontal profile
        # ax0 = plt.axes([0.10, 0.76, 0.85, 0.20])
        # ax1 = plt.axes([0.10, 0.10, 0.85, 0.65])
        # ax
        #
        # # Show uniformity image on top and fill whole image
        # ax0.imshow(im, cmap='gray', aspect='auto')
        #
        # # Make the tick labels invisible
        # plt.setp(ax0.get_xticklabels(), visible=False)
        # plt.setp(ax0.get_yticklabels(), visible=False)
        #
        # # Show profile below
        # ax1.plot(x_values, profile, label="profile")
        #
        # # Determine x_max & y_max
        # ymax = np.max(profile) + 0.1
        # xmax = np.max(x_values)
        #
        # # Set the x and y labels
        # ax1.set_xlabel("position [px]")
        # ax1.set_ylabel("signal [a.u.]")
        # ax1.grid(which='major', axis='y')
        # ax1.legend()
        #
        # if profile_type == 'horizontal':
        #     # Plot the weak, dead and mean limits and plot the buckets
        #     ax1.plot([x_values[0], x_values[-1]], [mean, mean], linestyle=':', linewidth=2, color='green', label="average")
        #     ax1.plot([x_values[0], x_values[-1]], [weak, weak], linestyle=':', linewidth=2, color='orange', label="weak")
        #     ax1.plot([x_values[0], x_values[-1]], [dead, dead], linestyle=':', linewidth=2, color='red', label="dead")
        #
        #     # Plot the buckets
        #     ax1.axvspan(x_values[0], x_values[pixels10], facecolor='black', alpha=0.2)      # Outer left 10 percent
        #     ax1.axvspan(x_values[-1], x_values[-pixels10], facecolor='black', alpha=0.2)    # Outer right 10 percent
        #     ax1.axvspan(x_values[pixels10], x_values[pixels30], facecolor='black', alpha=0.1) # Outer left 10-30 percent
        #     ax1.axvspan(x_values[-pixels30], x_values[-pixels10], facecolor='black', alpha=0.1) #Outer right 10-30 percent
        #     ax1.set_ylim(bottom=0, top=ymax)
        #
        #     # Offset for xlimit to see the first and last lines
        #     xoffset = 5
        #     ax1.set_xlim(left=-xoffset, right=xmax + xoffset)
        #     ax0.set_xlim(left=-xoffset, right=xmax + xoffset)
        #
        #     self.fig_uni = fig
        #
        #     # Save image
        #     fig.savefig(f"{self.path_save_images}/{self.label}uniformity.png", bbox_inches='tight')
        #
        # else:
        #     # Set limits for x and y axis
        #     ax0.set_xlim(left=0, right=xmax)
        #     ax1.set_ylim(bottom=0, top=ymax)
        #     ax1.set_xlim(left=0, right=xmax)
        #
        #     # Save image
        #     fig.savefig(f"{self.path_save_images}/{self.label}sensitivity.png", bbox_inches='tight')

    def run(self):
        """
        Wrapper to start all analysis steps:
        1. Isolate reverberation pattern in image
        2. Check if data is curved, and if so calculate limiting angles and radii of curvature
        3.  a) if data is curved: transform the curved image to a rectangular image
            b) if data is not curved: crop te data to a box containing the reverberation pattern
        4. Perform uniformity analysis: find response per crystal
        5. Perform sensitivity analysis: find maximum pPerenetration depth
        6. Save report containing the results to excel
        """

        # Step 1: Isolate the transducer mask and the reverberation pattern
        self.isolate_reverberation()
        transducer_mask = self.transducer_mask
        reverberation_mask = self.reverb_mask

        # Step 2: Check if data is curved, and if curved calculate limiting angles and radii
        self.is_curved = self.check_if_curved(reverberation_mask, type = 'reverberation')
        curv_trans = self.check_if_curved(transducer_mask, type='transducer')

        # Step 3a: If data is curved transform the curved pattern to a rectangle pattern
        if self.is_curved == True:
            reverb_uniform_data = self.curve_straightening(reverberation_mask, type = 'reverberation')
            transducer_data = self.curve_straightening(transducer_mask, type='transducer')

        # Step 3b: If data is not curved: crop te data to a box containing the reverberation pattern
        if self.is_curved == False:
            reverb_uniform_data = self.crop_image(reverberation_mask, type = 'reverberation')
            transducer_data = self.crop_image(transducer_mask, type='transducer')

        # Step 4: Perform uniformity analysis of the reverberation pattern: find response per crystal
        self.uniformity(reverb_uniform_data)

        # Step 5: Perform sensitivity analysis: find maximum penetration depth
        sensitivity_data = self.crop_sens_image(transducer_data)   # Take only middle 33% of the image for the sensitivity analysis
        self.sensitivity(sensitivity_data)

        # Step 6: Save report containing the results to excel
        self.export_results()

        # Step 7: Create final figure



#%%
qc = air_analysis('settings_data1.yaml', 'meta_data.yaml')
qc.run()

#%%
init = initialize_data('settings_data1.yaml', 'meta_data.yaml')
init.run()
#%%
fig_uni = qc.fig_uni
fig_sens = qc.fig_sens
fig_total = plt.figure()
fig_sens.show()

#
