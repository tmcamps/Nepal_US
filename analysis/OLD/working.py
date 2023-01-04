#%% Import packages
import os
import yaml
import pydicom
import numpy as np

import scipy.ndimage as scind
from PIL import Image
import operator
from scipy import optimize

import matplotlib.pyplot as plt


path = os.getcwd()                                                      # Get the current working directory of a process
settings_file = '../settings/meta_data.yaml'  # Name of file with preset settings
path_settings = os.path.join(path, settings_file)                       # Get directory of settings file
meta_data = yaml.load(open(path_settings), Loader=yaml.FullLoader)   # Load user settings from yaml file as dictionary


#%% Set runtime parameters
params = {'auto_suffix': False,
          'circle_fitfrac': 1. / 3,
          'cluster_fminsize': 10 * 10 * 3,
          'cluster_mode': "all_middle",
          'f_dead': 0.3,
          'f_weak': 0.5,
          'vcor_px': 0,
          'hcor_px': 0,
          'init_pt_x0y0x1y1': None,
          'pt_x0y0x1y1': None,
          'signal_thresh': 0,
          'pt_curve_radii_px': None,
          'pt_curve_origin_px': None,
          'pt_curve_angles_deg': None}

#%% Prepare input

path = '/analysis\\'  # Get the current working directory of a process
settings_file = 'settings_data2.yaml'                                         # Name of file with preset settings
path_settings = os.path.join(path, settings_file)                       # Get directory of settings file
settings_dct = yaml.load(open(path_settings), Loader=yaml.FullLoader)   # Load user settings from yaml file as dictionary

path_data = os.path.join(path, settings_dct['path_data'])               # Load directory for data
path_save = os.path.join(path, settings_dct['path_save'])               # Load directory for results
filenames = os.listdir(path_data)
data_file = os.path.join(path_data, filenames[0])

data = pydicom.dcmread(data_file, force=True)

# Set filename for save
date = data[0x00080020].value
transducer_type = data.TransducerType

fname = transducer_type + '-' + date + '-'
# Transpose the pixel array
# im = data.pixel_array.transpose()
# im_transpose = np.transpose(im, (1,0))

# Pixel array
im = data.pixel_array

#%% Initialize report dictionary
report = {}

#%% Isolate the reverberation pattern
"""
1. Restrict to bbox if provided
2. Find reverberations as largest connected component
3. Return reverb mask
"""
# Initialize report section for pattern
report_section = 'pattern'
report[report_section] = {}

#1. Create binary image with provided signal threshold
signal_thresh = params['signal_thresh']
BW = im > signal_thresh

#2. Use bbox if provided? Not necessary

#3. Determine components and their labels
cca, nb_labels = scind.label(BW)

#4. Select all clusters present in vertical middle area of image, excluding top and 0
h, w = np.shape(cca)

# Remove very small clusters
minsize = h * w * params['cluster_fminsize']

# Calculate the clustersizes
clustersize = scind.sum(BW, cca, range(nb_labels + 1))

# Check minimal size with the clustersize
while sum(clustersize > minsize) < 2 and minsize > 100:
    minsize = int(minsize / 10)

# Remove small clusters
mask_size = clustersize < minsize           # Determine the small clusters
cca[mask_size[cca]] = 0                     # Set small cluster to 0
labels = np.unique(cca)
nb_labels = len(labels)
cca = np.searchsorted(labels, cca)          # Create new components array without small clusters

# Search for clusters present in vertical middle area of image
search_middle = cca[:, int(0.4*w):int(0.6*w)]

labels_middle = []
for ss in search_middle.ravel():
    if ss > 0 and not ss in labels_middle:           # Find label of components present in vertical middle
        labels_middle.append(ss)                     # Append to label

# # Exclude labels in top rows; not part of image data but full of annotations
# search_top = cca[0:5, :]
# notlabs = []
#
# for ss in search_top.ravel():
#     if ss > 0 and not ss in notlabs:           # Find label of components present in vertical middle
#         notlabs.append(ss)                     # Append to label

# Create reverberation mask of only largest cluster and cluster present in vertical middle
reverberation_mask = np.reshape(np.in1d(cca, labels_middle), np.shape(cca))

# Create list with indices of the cluster representing the reverberation pattern
clus = np.where(reverberation_mask)
clus = [(x, y) for x, y in zip(clus[0], clus[1])]

#%%
# Create and save image of only largest cluster
reverb_image = im * reverberation_mask
reverb_image = Image.fromarray(reverb_image)
reverb_image.save(f"{path_save}/{fname}image_reverberation.png")

#%%
# Determine upper and lower edges of the cluster
rev_minx = min(clus, key=operator.itemgetter(1))[1]
rev_maxx = max(clus, key=operator.itemgetter(1))[1]
rev_maxy = max(clus, key=operator.itemgetter(0))[0]
rev_miny = min(clus, key=operator.itemgetter(0))[0]

#%%
if params.get('pt_x0y0x1y1', None) is None:
    params['pt_x0y0x1y1'] = [rev_minx, rev_miny, rev_maxx, rev_maxy]

    report[report_section]['box_xmin_px'] = ('int', rev_minx)
    report[report_section]['box_ymin_px'] = ('int', rev_miny)
    report[report_section]['box_xmax_px'] = ('int', rev_maxx)
    report[report_section]['box_ymax_px'] = ('int', rev_maxy)


#%% Transform reverberation data to rectangle if needed
"""
1. Check if data is curved
2. If True => straighten curve
3. Else => Get rectangular slab of data
"""

#%% 1. Find curve
""" 
Fit circle through top of reverb image """
report_section = 'curved'
report[report_section] ={}
was_curved = True

# Find outer edges of the image
x0, y0, x1, y1 = params['pt_x0y0x1y1']

# Find center x value
center_x = int(0.5 * (x0 + x1) + .5)


# From top of reverb image down, look for pixels != 0 from mid to left and from mid to right
# If both found, add it to the list
circL_xy = []
circR_xy = []
for y in range(y0, y1 + 1):
    for x in reversed(range(x0, center_x)):  # find left point
        xl = -1
        xr = -1
        if reverberation_mask[y, x]:
            xl = x
            if xl > -1:
                circL_xy.append((xl, y))
            break
    for x in range(center_x, x1):  # find right point
        if reverberation_mask[y, x]:
            xr = x
            if xr > -1:
                circR_xy.append((xr, y))
            break

    if xr-xl < 10:
        break

#%% Create complete circle to fit image
circ_xy = []
circ_xy.extend(circL_xy)
circ_xy.extend(circR_xy)
circ_xy.sort(key=operator.itemgetter(1))
circle_fitfrac = params['circle_fitfrac']

#%% Determine if transducer was curved
if len(circ_xy) < 11:
    # at least 10 point, else probably not a curved probe
    was_curved = False

if was_curved:
    # Use only central part for fitting, as deformations towards edges occur
    if circle_fitfrac < 1 and circle_fitfrac > 0:
        fraction = 1 - circle_fitfrac
        circxy_fraction = circ_xy[int(fraction * len(circ_xy)):]
        x = [xy[0] for xy in circxy_fraction]
        y = [xy[1] for xy in circxy_fraction]
        residu = -1.

    # Fit circle to transducer image
    center_estimate = np.mean(x), np.mean(y)

    # # Calculate the distance of each 2D points from the center
    # dist_center = np.sqrt((x-center_estimate[0])**2 + (y-center_estimate[1])**2)

    # Calculate algebraic distance between data points and the mean circle center
    def d_mean(center_estimate):
        # Calculate the distance of each 2D points from the center
        dist_center = np.sqrt((x - center_estimate[0]) ** 2 + (y - center_estimate[1]) ** 2)
        return dist_center - dist_center.mean()

    # Optimize center estimation
    center_optimize, ier = optimize.leastsq(d_mean, center_estimate)

    # Calculate distance of each 2D points from the optimized centre
    dist_center2 = np.sqrt((x-center_optimize[0])**2 + (y-center_optimize[1])**2)
    radius2 = dist_center2.mean()               # Determine radius of the circle

    # Calculate new residu
    residu2 = np.sum((dist_center2 - radius2 ) **2 )

    # Calculate limiting angles and radii
    xc = center_optimize[0]
    yc = center_optimize[1]
    curve_angles = [np.arctan2(circL_xy[0][0] - xc, circL_xy[0][1] - yc),
                    np.arctan2(circR_xy[0][0] - xc, circR_xy[0][1] - yc)]
    maxrad = min([
        (x0 - xc) / np.sin(curve_angles[0]),
        (x1 - xc) / np.sin(curve_angles[1]),
        (y1 - yc) ])

    # Save the values
    params['pt_curve_radii_px'] = [radius2, maxrad]
    params['pt_curve_origin_px'] = [xc, yc, radius2]
    params['pt_curve_angles_deg'] = [c / np.pi * 180. for c in curve_angles]

    report[report_section]['curve_residu'] = ('float', residu2)
    report[report_section]['curve_radmin_px'] = ('float', params['pt_curve_radii_px'][0])
    report[report_section]['curve_radmax_px'] = ('float', params['pt_curve_radii_px'][1])
    report[report_section]['curve_xc_px'] = ('float', xc)
    report[report_section]['curve_yc_px'] = ('float', yc)
    report[report_section]['curve_Rc_px'] = ('float', radius2)
    report[report_section]['curve_angmin_deg'] = ('float', params['pt_curve_angles_deg'][0])
    report[report_section]['curve_angmax_deg'] = ('float', params['pt_curve_angles_deg'][1])

#%% Transform data if reverb data is curved
if was_curved:
    """
    transform reverb pattern to rectangle: interpolate at coords
    """

    im_segment = im * reverberation_mask

    # Define necessary parameters
    x0, y0, x1, y1 = params['pt_x0y0x1y1']

    curve_angles = [c / 180. * np.pi for c in params['pt_curve_angles_deg']]      # Back to rad
    curve_radii = params['pt_curve_radii_px']                                # [Radius,maxrad]
    curve_xyr = params['pt_curve_origin_px']                                 # [xc,yc,radius]

    angles =  np.linspace(curve_angles[0], curve_angles[1], x1 - x0)
    rad = np.linspace(curve_radii[0], curve_radii[1], int(0.5 + curve_radii[1] - curve_radii[0]))
    an, ra = np.meshgrid(angles,rad)         # Return coordinate matrices from coordinate vectors.

    # Transform image to rectangular image using map_coordinates
    xi = curve_xyr[0] + ra * np.sin(an)
    yi = curve_xyr[1] + ra * np.cos(an)

    coordinates = np.array([yi,xi])
    rect_image = scind.map_coordinates(im_segment, coordinates)

    image_rect = Image.fromarray(rect_image)
    image_rect.save(f"{path_save}/{fname}image_uncurved.png")

else:
    """
     crop the data to a box containing the reverberation pattern,
     optionally removing some lines left/right and up/down
     """
    x0, y0, x1, y1 = params['pt_x0y0x1y1']
    dx = params['hcor_px']
    dy = params['vcor_px']
    rect_image = (im * reverberation_mask)[y0 + dy:y1 - dy + 1, x0 + dx:x1 - dx + 1]

    image_rect = Image.fromarray(rect_image)
    image_rect.save(f"{path_save}/{fname}image_cropped.png")

#%% Uniformity analysis; find response per crystal
"""
Get a rectangular slab of a fixed number of rings.
Average in vertical direction.
Count the number of pixel with response below the weak treshold, and the dead treshold.
Separate analysis for all, the outer 10%, outer 10%-30%, and 30%-70%
Also report the relative (to the overall mean) mean, min, max of all region
"""
report_section = "uniformity"
report[report_section] = {}

im_uni = rect_image

# Find maximum pixel value and safe to report
pix_max = int(np.max(im_uni) + .5)
report[report_section]['unif_max_pixval'] = ('int', pix_max)
# if pix_max > 253:
#     print("Warning! Saturated data. Use a lower gain!")

# Set values for weak element value and dead element value
weak = params['f_weak']
dead = params['f_dead']

# Average in vertical direction for horizontal profile
horizontal_profile = np.average(im_uni, axis=0)

# Report width of profile in pixels
if was_curved == False:
    report[report_section]['unif_box_width_px'] = ('int', len(horizontal_profile))
else:
    ang0, ang1 = params['pt_curve_angles_deg']
    report[report_section]['unif_curve_width_deg'] = ('float', ang1 - ang0)

# Report width of profile in mm
delta_x = data.SequenceOfUltrasoundRegions[0]['0x0018602c'].value  # physical delta X
unit = data.SequenceOfUltrasoundRegions[0]['0x00186024'].value  # Physical Units X & Y Direction

if unit == 3:
    length_mm = 10.*delta_x*len(horizontal_profile)
else:
    raise ValueError("Unit has to be set to mm")

report[report_section]['unif_width_mm']  = ('float', length_mm)

# Determine line profiles for mean, weak and dead
mean = np.average(horizontal_profile)
weak = weak * mean
dead = dead * mean

#%% Analysis of weak and dead elements
weak_idx = sorted(np.where(horizontal_profile < weak)[0])
dead_idx = sorted(np.where(horizontal_profile < dead)[0])

"""
count number of elements in list and number of neighboring elements
"""
def analyse_uniformity_idx(idxs):
    neighbors = 0
    if len(idxs) > 0:
        idxs = sorted(idxs)
        for i in range(len(idxs) - 1):
            if idxs[i + 1] == idxs[i] + 1:
                neighbors += 1

    return len(idxs), neighbors

#%% Analyse all and per bucket
num = len(horizontal_profile)
idx = list(range(num))
pct_10 = max(0, min(int(num * .1 + .5), num - 1))
pct_30 = max(0, min(int(num * .3 + .5), num - 1))
buckets = [
    ('all', idx),
    ('*00_10', idx[:pct_10] + idx[-pct_10:]),
    ('*10_30', idx[pct_10:pct_30] + idx[-pct_30:-pct_10]),
    ('10_90', idx[pct_10:-pct_10]),
    ('30_70', idx[pct_30:-pct_30]),
]

#%% Fill buckets
for lab, idx_valid in buckets:
    weak_num, weak_neighbors = analyse_uniformity_idx([i for i in weak_idx if i in idx_valid])
    dead_num, dead_neighbors = analyse_uniformity_idx([i for i in dead_idx if i in idx_valid])
    report[report_section]['unif_weak_{}'.format(lab)] = ('int', weak_num)
    report[report_section]['unif_weaknbs_{}'.format(lab)] = ('int', weak_neighbors)
    report[report_section]['unif_dead_{}'.format(lab)] = ('int', dead_num)
    report[report_section]['unif_deadnbs_{}'.format(lab)] = ('int', dead_neighbors)
    loc_prof = horizontal_profile[idx_valid]
    if not lab == 'all':
        report[report_section]['unif_relmean_{}'.format(lab)] = ('float', np.average(loc_prof) / mean)
        report[report_section]['unif_relmin_{}'.format(lab)] = ('float', np.min(loc_prof) / mean)
        report[report_section]['unif_relmax_{}'.format(lab)] = ('float', np.max(loc_prof) / mean)

# fill report with other results
report[report_section]['unif_mean'] = ('float', mean)
report[report_section]['unif_relmin'] = ('float', np.min(horizontal_profile) / mean)
report[report_section]['unif_relmax'] = ('float', np.max(horizontal_profile) / mean)
report[report_section]['unif_f_weak'] = ('float', params['f_weak'])
report[report_section]['unif_f_dead'] = ('float', params['f_dead'])

#%%
# stack ring pattern on top of profile image
pos = np.array(list(range(len(horizontal_profile))))
fig = plt.figure()
# plt.title("uniformity")
ax0 = plt.axes([0.10, 0.76, 0.85, 0.20])  # , adjustable='box', aspect=myDataAspect) #x0, y0, width, height
ax1 = plt.axes([0.10, 0.10, 0.85, 0.65])  # , adjustable='box', aspect=myDataAspect)

# show ring pattern on top; fill whole image
# ax0.axis('off')
ax0.imshow(im_uni, cmap='gray', aspect='auto')

# make these tick labels invisible
plt.setp(ax0.get_xticklabels(), visible=False)
plt.setp(ax0.get_yticklabels(), visible=False)

# show profile below
ax1.plot(pos, horizontal_profile, label="profile")
ax1.plot([pos[0], pos[-1]], [mean, mean], linestyle=':', linewidth=2, color='green', label="average")
ax1.plot([pos[0], pos[-1]], [weak, weak], linestyle=':', linewidth=2, color='orange', label="weak")
ax1.plot([pos[0], pos[-1]], [dead, dead], linestyle=':', linewidth=2, color='red', label="dead")

# add buckets
ymax = np.max(horizontal_profile) + 1
xmax = np.max(pos)
ax1.axvspan(pos[0], pos[pct_10], facecolor='black', alpha=0.2)
ax1.axvspan(pos[-1], pos[-pct_10], facecolor='black', alpha=0.2)
ax1.axvspan(pos[pct_10], pos[pct_30], facecolor='black', alpha=0.1)
ax1.axvspan(pos[-pct_30], pos[-pct_10], facecolor='black', alpha=0.1)
ax1.set_ylim(bottom=0, top=ymax)
# ax1.set_xlim(left=0, right=xmax)

offset = 5
ax1.set_xlim(left=-offset, right=xmax + offset)  # want to see first and last line
ax0.set_xlim(left=-offset, right=xmax + offset)  # want to see first and last line
ax1.set_xlabel("position [px]")
ax1.set_ylabel("signal [a.u.]")
ax1.grid(which='major', axis='y')
ax1.legend()


fig.savefig(f"{path_save}/{fname}uniformity.png", bbox_inches='tight')

#%% Perform sensitivity analysis
"""
data: crop the data to a box containing the reverberation pattern
then take only middle 33% for sensitivity analysis
"""
h, w = im_uni.shape

pct_x = max(0, min(int(w * .3 + .5), w - 1))
im_sens = im_uni[:, pct_x:-pct_x]

#%% Perform analysis
"""
Get a rectangular slab of the reverberation pattern.
Average in horizontal direction.
Just turn it into a nice image, and report the total depth of the pattern.
"""
report_section = "sensitivity"
report[report_section] = {}

# average profile horizontally
vertical_profile = np.average(im_sens, axis=1)
pos = np.array(list(range(len(vertical_profile))))


#%%
## make the image
# stack ring pattern on top of profile image
xmax = np.max(pos)

fig = plt.figure()
ax0 = plt.axes([0.10, 0.76, 0.85, 0.20])  # , adjustable='box', aspect=myDataAspect) #x0, y0, width, height
ax1 = plt.axes([0.10, 0.10, 0.85, 0.65])  # , adjustable='box', aspect=myDataAspect)

# show ring pattern on top; fill whole image
# ax0.axis('off')
ax0.imshow(np.transpose(im_sens, (1, 0)), cmap='gray', aspect='auto')
ax0.set_xlim(left=0, right=xmax)

# make these tick labels invisible
plt.setp(ax0.get_xticklabels(), visible=False)
plt.setp(ax0.get_yticklabels(), visible=False)

# show profile below
ax1.plot(pos, vertical_profile, label="profile")
ymax = np.max(vertical_profile) + 1

ax1.set_ylim(bottom=0, top=ymax)
ax1.set_xlim(left=0, right=xmax)
ax1.set_xlabel("position [px]")
ax1.set_ylabel("signal [a.u.]")
ax1.grid(which='major', axis='y')
ax1.legend()

fig.savefig(f"{path_save}/{fname}sensitivity.png", bbox_inches='tight')

#%% Save data to excel
from openpyxl import Workbook, load_workbook

if os.path.exists(f"{path_save}/results.xlsx"):
    wb = load_workbook(f"{path_save}/results.xlsx")
    if fname in wb.sheetnames:
        wb.remove(wb[fname])
        sheet = wb.create_sheet(title=fname)
    else:
        sheet = wb.create_sheet(title=fname)
else:
    wb = Workbook()
    sheet = wb.active
    sheet.title = fname

sheet["A1"] = "Section"
sheet["B1"] = "Variable"
sheet["C1"] = "Value"

start_row = 2
for section in report.keys():
    sheet[f"A{start_row}"] = section

    for row2, (key,vals) in enumerate(report[section].items()):
        row2 += start_row
        sheet[f"B{row2}"] = key
        sheet[f"C{row2}"] = vals[1]

    start_row += len(report[section])

wb.save(f"{path_save}/results.xlsx")

