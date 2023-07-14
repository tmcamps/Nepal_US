import util_export as util
import pandas as pd
import openpyxl
import os
from openpyxl.utils.dataframe import dataframe_to_rows
#%%

def create_dataframe(qc):
    dict = qc.report
    df = util.create_df(dict)
    df1 = util.create_multilevel(df)
    df2 = util.stack_dataframe(df1)
    df3 = util.add_date(df2, qc.date)

    return df3

def save_to_excel(qc):
    df = create_dataframe(qc)
    sheet_name = qc.label

    wb = openpyxl.load_workbook(f"{qc.path_save_overview}/results.xlsx")

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        sheet = wb.create_sheet(title=sheet_name)
    else:
        sheet = wb.create_sheet(title=sheet_name)

    for r in dataframe_to_rows(df, index=True, header=True):
        sheet.append(r)

    wb.save(f"{qc.path_save_overview}/results.xlsx")


